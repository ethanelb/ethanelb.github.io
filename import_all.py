#!/usr/bin/env python3
"""
Full import of 0R (Bobines) + 0S (Palettes) Excel files into Supabase.
Table columns: quality, color, details, gsm, width, weight, price, ref, noyau, format, image_url
"""

import openpyxl, re, json, math, time
from urllib.request import urlopen, Request
from urllib.error import HTTPError, URLError

SURL = 'https://bvcgpdoukhcatjibmvnb.supabase.co'
SKEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImJ2Y2dwZG91a2hjYXRqaWJtdm5iIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzIyNzg5MjgsImV4cCI6MjA4Nzg1NDkyOH0.Ip3ykSUS9sajTH04yXBerOG1haBKMD1kAvMQNjnGL1Q'

BASE_HEADERS = {
    'apikey': SKEY,
    'Authorization': f'Bearer {SKEY}',
    'Content-Type': 'application/json',
}

PHOTO_BASE = 'https://stock.prodi.net/albums/photo/'

def simplify_color(raw):
    """'BLANC / WHITE' → 'Blanc'"""
    if not raw:
        return None
    french = raw.split('/')[0].strip()
    if not french:
        return None
    return french.capitalize()

def parse_price(val):
    """'800Eur/T' → 800.0"""
    if not val:
        return None
    m = re.match(r'[\d.]+', str(val).strip())
    if m:
        p = float(m.group())
        return p if p > 0 else None
    return None

def to_int(val):
    if val is None:
        return None
    try:
        v = int(float(val))
        return v if v > 0 else None
    except (ValueError, TypeError):
        return None

def to_float(val):
    if val is None:
        return None
    try:
        v = float(val)
        return v if v > 0 else None
    except (ValueError, TypeError):
        return None

def parse_bobines():
    """Parse 0R - Bobines. Data starts row 28."""
    wb = openpyxl.load_workbook('/Users/tantan/Downloads/0R - STOCK DEPOT BOBINE.xlsx', data_only=True)
    ws = wb.active
    products = []
    for row in ws.iter_rows(min_row=28, values_only=True):
        ref_cell = row[0]
        if ref_cell is None:
            continue
        ref = str(ref_cell).strip()
        if not ref.startswith('Photo_'):
            continue

        photo_id = ref.replace('Photo_', '').strip()

        # Build details string from multiple sub-fields
        details_parts = []
        raw_details = str(row[3]).strip() if row[3] else ''
        if raw_details and raw_details not in ('None', ''):
            # Clean up the details (strip trailing dashes and spaces)
            cleaned = re.sub(r'\s*-\s*-\s*$', '', raw_details).strip()
            if cleaned:
                details_parts.append(cleaned)
        diam = to_int(row[6])
        if diam:
            details_parts.append(f'Ø{diam}mm')

        product = {
            'ref': ref,
            'quality': str(row[1]).strip() if row[1] else None,
            'color': simplify_color(str(row[2]).strip() if row[2] else None),
            'details': ' · '.join(details_parts) if details_parts else None,
            'gsm': to_int(row[4]),
            'width': to_int(row[5]),
            'weight': to_float(row[8]),
            'price': parse_price(row[9]),
            'noyau': str(to_int(row[7])) if to_int(row[7]) else None,
            'longueur': None,
            'format': 'Bobine',
            'image_url': f'{PHOTO_BASE}{photo_id}.jpg',
        }
        products.append(product)

    print(f"Parsed {len(products)} bobines")
    return products

def parse_palettes():
    """Parse 0S - Palettes. Data starts row 32."""
    wb = openpyxl.load_workbook('/Users/tantan/Downloads/0S - STOCK DEPOT PALETTE REUNIS EXTERIEURS ET DEPOT.xlsx', data_only=True)
    ws = wb.active
    products = []
    for row in ws.iter_rows(min_row=32, values_only=True):
        ref_cell = row[0]
        if ref_cell is None:
            continue
        ref = str(ref_cell).strip()
        if not ref.startswith('Photo_'):
            continue

        photo_id = ref.replace('Photo_', '').strip()

        # Details: row[3] = description only; longueur stored as proper column
        details_parts = []
        raw_details = str(row[3]).strip() if row[3] else ''
        if raw_details and raw_details not in ('None', ''):
            cleaned = re.sub(r'\s*-\s*-\s*$', '', raw_details).strip()
            if cleaned:
                details_parts.append(cleaned)
        longueur = to_int(row[6])

        product = {
            'ref': ref,
            'quality': str(row[1]).strip() if row[1] else None,
            'color': simplify_color(str(row[2]).strip() if row[2] else None),
            'details': ' · '.join(details_parts) if details_parts else None,
            'gsm': to_int(row[4]),
            'width': to_int(row[5]),
            'longueur': longueur,
            'weight': to_float(row[7]),
            'price': parse_price(row[8]),
            'noyau': None,
            'format': 'Palette',
            'image_url': f'{PHOTO_BASE}{photo_id}.jpg',
        }
        products.append(product)

    print(f"Parsed {len(products)} palettes")
    return products

def api_request(method, path, body=None, extra_headers=None):
    """Make a Supabase REST API request. Returns (status, body_str)."""
    url = f'{SURL}/rest/v1/{path}'
    headers = {**BASE_HEADERS}
    if extra_headers:
        headers.update(extra_headers)
    data = json.dumps(body).encode('utf-8') if body is not None else None
    req = Request(url, data=data, headers=headers, method=method)
    try:
        with urlopen(req, timeout=30) as resp:
            return resp.status, resp.read().decode('utf-8')
    except HTTPError as e:
        return e.code, e.read().decode('utf-8')
    except URLError as e:
        return 0, str(e)

def count_products():
    url = f'{SURL}/rest/v1/products?select=id'
    req = Request(url, headers={**BASE_HEADERS, 'Prefer': 'count=exact', 'Range': '0-0'})
    try:
        with urlopen(req, timeout=10) as resp:
            cr = resp.getheader('Content-Range', '')
            if '/' in cr:
                return int(cr.split('/')[-1])
    except Exception:
        pass
    return -1

def delete_all():
    """Delete all rows from products table."""
    print("Deleting all existing products...")
    status, body = api_request('DELETE', 'products?id=gte.0')
    if status in (200, 204):
        print("  Deleted OK")
    else:
        print(f"  Delete status {status}: {body[:200]}")

def insert_batch(rows, idx, total):
    """Insert a batch. Returns (ok, error_str)."""
    status, body = api_request('POST', 'products', rows)
    if status in (200, 201):
        return True, None
    return False, f'HTTP {status}: {body[:300]}'

def main():
    print("=== Prodiconseil — Full Import ===\n")

    existing = count_products()
    print(f"Current products in DB: {existing}")

    bobines = parse_bobines()
    palettes = parse_palettes()
    all_products = bobines + palettes
    print(f"Total products to import: {len(all_products)}")

    # Clean slate
    if existing > 0:
        delete_all()

    # Insert in batches of 400
    BATCH = 400
    total_batches = math.ceil(len(all_products) / BATCH)
    ok_count = 0
    errors = []

    print(f"\nInserting {total_batches} batches of {BATCH}...")
    for i in range(total_batches):
        batch = all_products[i*BATCH:(i+1)*BATCH]
        ok, err = insert_batch(batch, i+1, total_batches)
        if ok:
            ok_count += len(batch)
            pct = int(ok_count / len(all_products) * 100)
            print(f"  [{pct:3d}%] Batch {i+1}/{total_batches} ✓  ({ok_count}/{len(all_products)})")
        else:
            errors.append(f"Batch {i+1}: {err}")
            print(f"  Batch {i+1}/{total_batches} ✗  {err}")
        if i < total_batches - 1:
            time.sleep(0.25)

    print(f"\n=== Done ===")
    print(f"Inserted: {ok_count}/{len(all_products)}")
    if errors:
        print(f"\nErrors ({len(errors)}):")
        for e in errors:
            print(f"  {e}")

    time.sleep(1)
    final = count_products()
    print(f"Products in DB now: {final}")

if __name__ == '__main__':
    main()
